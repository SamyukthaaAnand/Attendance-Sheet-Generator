import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.worksheet.page import PageMargins
from flask import Flask, request, send_file, render_template
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if request.method == 'POST':
        # Get the file from the request
        file = request.files['file']
        
        # Get the selected subject and batch size from the form
        subject = request.form['subject'].strip().lower()
        batch_size = int(request.form['batch_size'])
        year = request.form['year'].strip().lower()
       
        subject_mapping = {
            # TE Subjects DLO/ILO
            "network programming": "NP",
            "data science fundamentals": "DSF",
            "artificial intelligence": "AI",
            "cryptography": "cryptography",
            "web development": "WD",
            "fundamentals of machine learning": "FML",
            "data warehousing and mining": "DWM",
            "cg vr & ar":"CGVR",
            "ai for business applications":"AIBA",
            "audio processing":"AP",
            "mechatronics":"mechatronics",
            "theory of automata and formal languages":"automata",
            "project management":"PM",
            "entrepreneurship development management":"EDM",
            "product lifecycle management":"PLM",
            "information security": "IS",
            "cloud computing": "CC",
            "blockchain": "BC",
            "internet of things": "IoT",
            "cyber security": "CS",
            "distributed systems": "DSY",
            "deep learning": "DL"
            # SE Subjects Major/Minor
        }

        
        short_subject = subject_mapping.get(subject, subject)  # Use the original subject name if not found in mapping

        # Read the Excel file
        df = pd.read_excel(file)

        # Define the custom batch order
        batch_order = ['A1', 'A2', 'A3', 'B1', 'B2', 'B3', 'C1', 'C2', 'C3']

        # Determine column names based on the selected year
        if year == 'te':
            subject_columns = ['DLO1', 'DLO2', 'ILO1', 'ILO2']
        else:
            subject_columns = ['major', 'minor']

        # Convert subject columns to lower case for case-insensitive matching
        for col in subject_columns:
            df[col] = df[col].str.lower()

        # Filter rows where the selected subject is in any of the subject columns
        filtered_df = df[df[subject_columns].apply(lambda x: subject in x.values, axis=1)]

        # Sort the DataFrame based on 'Batch' column according to the custom batch order
        filtered_df['Batch'] = pd.Categorical(filtered_df['Batch'], categories=batch_order, ordered=True)
        filtered_df = filtered_df.sort_values(by='Batch')

        # Prepare the output file name using the short form of the subject
        output_file = f'{short_subject}.xlsx'
        
        # Create the Excel file
        wb = Workbook()
        
        # Create a thin border style for all sides
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Generate multiple sheets if the number of students exceeds batch size
        num_batches = (len(filtered_df) + batch_size - 1) // batch_size
        for i in range(num_batches):
            start_idx = i * batch_size
            end_idx = min(start_idx + batch_size, len(filtered_df))
            batch_df = filtered_df.iloc[start_idx:end_idx]

            # Create a new sheet for each batch, named short_subject1, short_subject2, etc.
            ws = wb.create_sheet(title=f'{short_subject}{i+1}')
            
            # Set headers and format for each sheet
            bold_font = Font(bold=True)

            # Helper function to apply border to merged cells
            def apply_border_to_merged_cells(ws, cell_range, border):
                cells = ws[cell_range]
                for row in cells:
                    for cell in row:
                        cell.border = border

            # Set up header rows with merged cells and apply borders
            ws.merge_cells('A1:C1')
            ws['A1'] = f"Batch : "
            ws['A1'].alignment = Alignment(horizontal='left')
            ws['A1'].font = bold_font
            apply_border_to_merged_cells(ws, 'A1:C1', thin_border)

            ws.merge_cells('A2:C2')
            ws['A2'] = f"Subject : "
            ws['A2'].alignment = Alignment(horizontal='left')
            ws['A2'].font = bold_font
            apply_border_to_merged_cells(ws, 'A2:C2', thin_border)

            ws.merge_cells('A3:C3')
            ws['A3'] = "Name of Faculty : "
            ws['A3'].alignment = Alignment(horizontal='left')
            ws['A3'].font = bold_font
            apply_border_to_merged_cells(ws, 'A3:C3', thin_border)

            ws.merge_cells('D1:W1')
            ws['D1'] = "Ramrao Adik Institute of Technology"
            ws['D1'].alignment = Alignment(horizontal='center')
            ws['D1'].font = bold_font
            apply_border_to_merged_cells(ws, 'D1:W1', thin_border)

            ws.merge_cells('D2:W2')
            ws['D2'] = "D Y Patil Deemed to be University"
            ws['D2'].alignment = Alignment(horizontal='center')
            ws['D2'].font = bold_font
            apply_border_to_merged_cells(ws, 'D2:W2', thin_border)

            ws.merge_cells('D3:W3')
            ws['D3'] = "Attendance Sheet of ODD Sem 2024-25      T.E.IT"
            ws['D3'].alignment = Alignment(horizontal='center')
            ws['D3'].font = bold_font
            apply_border_to_merged_cells(ws, 'D3:W3', thin_border)

            # Adding the column headers: "Sr. No", "Batch", "Roll No", "Name" followed by 19 columns (numbered 1 to 19)
            headers = ["Sr. No", "Batch", "Roll No", "Name"] + list(range(1, 20))
            ws.append(headers)
            
            # Set widths for the columns
            ws.column_dimensions['A'].width = 5  # Reduced width for "Sr. No"
            ws.column_dimensions['B'].width = 10  # Reduced width for "Batch"
            ws.column_dimensions['D'].width = 20  # Wider width for "Name"
            
            # Set width for numbered columns (1 to 19) to a smaller size (e.g., 5)
            for col in range(5, 24):  # Columns E to W correspond to numbers 1 to 19
                ws.column_dimensions[chr(64 + col)].width = 5  # Small width for number columns

            # Add borders to the header row
            for row in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=23):
                for cell in row:
                    cell.border = thin_border

            # Reset serial number for each batch
            serial_number = 1

            # Adding student data to the sheet
            for idx, row in batch_df.iterrows():
                # Create a row starting with the serial number, batch, roll number, name and then 19 empty cells
                data_row = [serial_number, row['Batch'], row['Roll No'], row['Name']] + [None] * 19
                ws.append(data_row)
                serial_number += 1  # Increment serial number for each student in the batch

            # Add borders to the data rows
            for row in ws.iter_rows(min_row=5, max_row=5 + len(batch_df), min_col=1, max_col=23):
                for cell in row:
                    cell.border = thin_border

            # Fit sheet into one landscape page
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

            # Set margins to adjust content
            ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

        # Save the Excel file
        del wb['Sheet']  # Remove the default sheet created by openpyxl
        wb.save(output_file)

        # Send the generated file to the user
        return send_file(output_file, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
